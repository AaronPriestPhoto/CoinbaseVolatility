#!/usr/bin/env python3
import argparse
import csv
import os
import sys
import time
import webbrowser
from datetime import datetime, timedelta, timezone

import numpy as np
import requests
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ----------------------------
# Self-contained script setup
# ----------------------------
# Ensure script runs from its own directory (useful for StreamDeck, etc.)
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

# ----------------------------
# Config
# ----------------------------
BASE_URL = "https://api.exchange.coinbase.com"
HEADERS = {"User-Agent": "Mozilla/5.0"}
GRANULARITY = 86400  # daily candles
RATE_LIMIT_SLEEP = 0.35  # ~3 requests/sec safety

# ----------------------------
# Date window: will be calculated in main() based on --days parameter
# ----------------------------


# ----------------------------
# Helpers
# ----------------------------
def iso_format(dt: datetime) -> str:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


def calculate_percentage_change(low: float, high: float) -> float:
    if low <= 0:
        return 0.0
    return (high - low) / low * 100.0


def safe_file_operation(operation_func, file_path: str, operation_name: str, max_retries: int = 10):
    """
    Safely perform file operations with user prompts when file is in use.
    
    Args:
        operation_func: Function to execute (should return True on success)
        file_path: Path to the file
        operation_name: Description of the operation for user prompts
        max_retries: Maximum number of retry attempts
    
    Returns:
        True if operation succeeded, False if user cancelled
    """
    for attempt in range(max_retries):
        try:
            return operation_func()
        except (PermissionError, OSError) as e:
            if attempt < max_retries - 1:
                print(f"\n⚠️  Cannot {operation_name} '{file_path}' - file may be open in another program.")
                print(f"   Error: {e}")
                print(f"\nPlease close the file and press Enter to retry, or type 'q' and Enter to quit:")
                
                user_input = input().strip().lower()
                if user_input == 'q':
                    print("Operation cancelled by user.")
                    return False
                print("Retrying...")
            else:
                print(f"\n❌ Failed to {operation_name} '{file_path}' after {max_retries} attempts.")
                print("Please close the file manually and run the script again.")
                return False
        except KeyboardInterrupt:
            print(f"\n\nOperation cancelled by user (Ctrl+C) during {operation_name}.")
            return False
        except Exception as e:
            print(f"❌ Unexpected error during {operation_name}: {e}")
            return False
    
    return False


def safe_remove_file(file_path: str) -> bool:
    """Safely remove a file with user prompts if it's in use."""
    if not os.path.exists(file_path):
        return True
    
    def remove_operation():
        os.remove(file_path)
        return True
    
    return safe_file_operation(remove_operation, file_path, "remove file")


def safe_write_file(file_path: str, write_func, operation_name: str) -> bool:
    """Safely write to a file with user prompts if it's in use."""
    def write_operation():
        write_func(file_path)
        return True
    
    return safe_file_operation(write_operation, file_path, operation_name)


def open_file(file_path: str) -> bool:
    """Open file with the default application."""
    try:
        # Convert to absolute path for better compatibility
        abs_path = os.path.abspath(file_path)
        
        # Use webbrowser to open files (works cross-platform)
        webbrowser.open(f"file://{abs_path}")
        return True
    except Exception as e:
        print(f"Warning: Could not auto-open file: {e}")
        return False


def create_excel_file(output_file: str) -> Workbook:
    """Create a new Excel workbook with headers and formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Volatility Analysis"
    
    # Add headers
    headers = ["Pair", "Volatility", "Volume", "MinFunds", "AvgLong%", "MaxLong%", "AvgShort%", "MaxShort%", "Sessions"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    return wb


def save_to_excel(pair: str, median_pct_change: float, volume: float, min_order_size, supertrend_stats: dict, wb: Workbook) -> None:
    """Add data row to Excel workbook."""
    ws = wb.active
    row = ws.max_row + 1
    
    # Add data as numeric values (Excel will handle formatting)
    ws.cell(row=row, column=1, value=pair)
    ws.cell(row=row, column=2, value=median_pct_change)
    ws.cell(row=row, column=3, value=int(volume) if volume > 0 else 0)
    
    # Ensure MinFunds is stored as a number
    if min_order_size == "N/A":
        ws.cell(row=row, column=4, value=0)
    else:
        try:
            # Convert to float to ensure it's a number
            ws.cell(row=row, column=4, value=float(min_order_size))
        except (ValueError, TypeError):
            ws.cell(row=row, column=4, value=0)
    
    # Add SuperTrend data
    ws.cell(row=row, column=5, value=supertrend_stats['avg_long_session_pct'])
    ws.cell(row=row, column=6, value=supertrend_stats['max_long_session_pct'])
    ws.cell(row=row, column=7, value=supertrend_stats['avg_short_session_pct'])
    ws.cell(row=row, column=8, value=supertrend_stats['max_short_session_pct'])
    ws.cell(row=row, column=9, value=supertrend_stats['total_sessions'])


def format_excel_file(wb: Workbook) -> None:
    """Format Excel file with auto-sized columns and styling."""
    ws = wb.active
    
    # Set specific column widths for better appearance
    # Pair column (A) - auto-size based on content
    max_pair_length = 0
    for row in range(1, ws.max_row + 1):
        cell_value = str(ws.cell(row=row, column=1).value or "")
        if len(cell_value) > max_pair_length:
            max_pair_length = len(cell_value)
    ws.column_dimensions['A'].width = min(max_pair_length + 2, 20)
    
    # Volatility column (B) - fixed width for 2 decimal places
    ws.column_dimensions['B'].width = 12
    
    # Volume column (C) - wider for large numbers with commas
    max_volume_length = 0
    for row in range(1, ws.max_row + 1):
        cell_value = str(ws.cell(row=row, column=3).value or "")
        if len(cell_value) > max_volume_length:
            max_volume_length = len(cell_value)
    ws.column_dimensions['C'].width = min(max_volume_length + 3, 25)
    
    # MinFunds column (D) - fixed width for currency
    ws.column_dimensions['D'].width = 12
    
    # SuperTrend columns (E-I) - fixed widths for percentages and sessions
    ws.column_dimensions['E'].width = 12  # AvgLong%
    ws.column_dimensions['F'].width = 12  # MaxLong%
    ws.column_dimensions['G'].width = 12  # AvgShort%
    ws.column_dimensions['H'].width = 12  # MaxShort%
    ws.column_dimensions['I'].width = 10  # Sessions
    
    # Format specific columns
    # Volatility column (B) - 2 decimal places
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        cell.number_format = '0.00'
    
    # Volume column (C) - Number format with commas (no decimals)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        cell.number_format = '#,##0'
    
    # MinFunds column (D) - Simple number format to avoid green triangles
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=4)
        # Use simple number format without currency symbol
        cell.number_format = '0.00'
    
    # SuperTrend columns (E-H) - 2 decimal places for percentages
    for col in range(5, 9):  # Columns E through H
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = '0.00'
    
    # Sessions column (I) - Integer format
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=9)
        cell.number_format = '0'


def sort_excel_by_volatility(wb: Workbook) -> None:
    """Sort Excel data by volatility (highest first)."""
    ws = wb.active
    
    # Get all data rows (excluding header)
    data_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, 5):  # 4 columns
            row_data.append(ws.cell(row=row, column=col).value)
        data_rows.append(row_data)
    
    # Sort by volatility (column 1) in descending order
    data_rows.sort(key=lambda x: float(x[1]) if x[1] is not None else 0, reverse=True)
    
    # Clear data rows and rewrite sorted data
    for row in range(2, ws.max_row + 1):
        for col in range(1, 5):
            ws.cell(row=row, column=col, value="")
    
    # Write sorted data
    for i, row_data in enumerate(data_rows, start=2):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)


def save_to_csv(pair: str, median_pct_change: float, volume: float, min_order_size, supertrend_stats: dict, output_file: str) -> bool:
    """Save data to CSV file with safe file handling. Returns True if successful, False if cancelled."""
    header = ["Pair", "Volatility", "Volume", "MinFunds", "AvgLong%", "MaxLong%", "AvgShort%", "MaxShort%", "Sessions"]
    file_exists = os.path.exists(output_file)
    
    def write_operation():
        with open(output_file, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(header)
            # Format: Volatility to 2 decimal places, Volume as integer with commas
            formatted_volume = f"{int(volume):,}" if volume > 0 else "0"
            writer.writerow([
                pair, 
                f"{median_pct_change:.2f}", 
                formatted_volume, 
                min_order_size,
                f"{supertrend_stats['avg_long_session_pct']:.2f}",
                f"{supertrend_stats['max_long_session_pct']:.2f}",
                f"{supertrend_stats['avg_short_session_pct']:.2f}",
                f"{supertrend_stats['max_short_session_pct']:.2f}",
                supertrend_stats['total_sessions']
            ])
        return True
    
    return safe_file_operation(write_operation, output_file, "write to CSV file")


def sort_csv_by_median(output_file: str) -> bool:
    """Sort CSV file by volatility with safe file handling. Returns True if successful, False if cancelled."""
    if not os.path.exists(output_file):
        return True
    
    def sort_operation():
        with open(output_file, "r", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        if not rows:
            return True
        header, data = rows[0], rows[1:]
        # Sort by volatility (column 1), parsing the formatted string back to float
        data.sort(key=lambda r: float(r[1]), reverse=True)
        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(data)
        return True
    
    return safe_file_operation(sort_operation, output_file, "sort CSV file")


# ----------------------------
# SuperTrend Analysis
# ----------------------------
def calculate_atr(high, low, close, length=10):
    """Calculate Average True Range (ATR) for SuperTrend calculation."""
    if len(high) < length + 1:
        return None
    
    tr_list = []
    for i in range(1, len(high)):
        tr1 = high[i] - low[i]
        tr2 = abs(high[i] - close[i-1])
        tr3 = abs(low[i] - close[i-1])
        tr_list.append(max(tr1, tr2, tr3))
    
    # Calculate ATR using simple moving average
    atr_values = []
    for i in range(length-1, len(tr_list)):
        atr = sum(tr_list[i-length+1:i+1]) / length
        atr_values.append(atr)
    
    return atr_values


def calculate_supertrend(high, low, close, factor=3, atr_length=10):
    """
    Calculate SuperTrend indicator.
    
    Args:
        high, low, close: Price arrays
        factor: SuperTrend factor (default: 3)
        atr_length: ATR period (default: 10)
    
    Returns:
        tuple: (supertrend_line, trend_direction, signals)
            - supertrend_line: SuperTrend line values
            - trend_direction: 1 for uptrend, -1 for downtrend
            - signals: 1 for buy signal, -1 for sell signal, 0 for no signal
    """
    if len(high) < atr_length + 1:
        return None, None, None
    
    # Calculate ATR
    atr_values = calculate_atr(high, low, close, atr_length)
    if not atr_values:
        return None, None, None
    
    # Initialize arrays
    supertrend_line = [None] * len(close)
    trend_direction = [0] * len(close)
    signals = [0] * len(close)
    
    # Calculate basic upper and lower bands
    basic_upper = []
    basic_lower = []
    
    for i in range(atr_length, len(close)):
        atr_idx = i - atr_length
        basic_upper.append((high[i] + low[i]) / 2 + factor * atr_values[atr_idx])
        basic_lower.append((high[i] + low[i]) / 2 - factor * atr_values[atr_idx])
    
    # Calculate final SuperTrend line
    for i in range(atr_length, len(close)):
        basic_upper_idx = i - atr_length
        basic_lower_idx = i - atr_length
        
        # Initial trend direction
        if i == atr_length:
            if close[i] <= basic_lower[basic_lower_idx]:
                trend_direction[i] = -1
                supertrend_line[i] = basic_lower[basic_lower_idx]
            else:
                trend_direction[i] = 1
                supertrend_line[i] = basic_upper[basic_upper_idx]
        else:
            prev_trend = trend_direction[i-1]
            
            if prev_trend == 1:
                # Previous trend was up
                if basic_lower[basic_lower_idx] > supertrend_line[i-1]:
                    supertrend_line[i] = basic_lower[basic_lower_idx]
                else:
                    supertrend_line[i] = supertrend_line[i-1]
            else:
                # Previous trend was down
                if basic_upper[basic_upper_idx] < supertrend_line[i-1]:
                    supertrend_line[i] = basic_upper[basic_upper_idx]
                else:
                    supertrend_line[i] = supertrend_line[i-1]
            
            # Determine current trend direction
            if close[i] > supertrend_line[i]:
                trend_direction[i] = 1
            else:
                trend_direction[i] = -1
            
            # Generate signals
            if trend_direction[i] != trend_direction[i-1]:
                if trend_direction[i] == 1:
                    signals[i] = 1  # Buy signal
                else:
                    signals[i] = -1  # Sell signal
    
    return supertrend_line, trend_direction, signals


def analyze_supertrend_sessions(high, low, close, supertrend_line, trend_direction, signals):
    """
    Analyze SuperTrend sessions to calculate % changes.
    
    Returns:
        dict: {
            'avg_long_session_pct': average % rise in long sessions,
            'max_long_session_pct': maximum % rise in any long session,
            'avg_short_session_pct': average % fall in short sessions,
            'max_short_session_pct': maximum % fall in any short session,
            'total_sessions': total number of sessions,
            'long_sessions': number of long sessions,
            'short_sessions': number of short sessions
        }
    """
    if not signals or not any(signals):
        return {
            'avg_long_session_pct': 0.0,
            'max_long_session_pct': 0.0,
            'avg_short_session_pct': 0.0,
            'max_short_session_pct': 0.0,
            'total_sessions': 0,
            'long_sessions': 0,
            'short_sessions': 0
        }
    
    long_session_changes = []
    short_session_changes = []
    current_session_start = None
    current_session_type = None
    current_session_high = None
    current_session_low = None
    
    for i, signal in enumerate(signals):
        if signal != 0:  # New session starts
            # Close previous session if exists
            if current_session_start is not None and current_session_type is not None:
                if current_session_type == 1:  # Long session
                    if current_session_high is not None:
                        session_change = ((current_session_high - close[current_session_start]) / close[current_session_start]) * 100
                        long_session_changes.append(session_change)
                else:  # Short session
                    if current_session_low is not None:
                        session_change = ((close[current_session_start] - current_session_low) / close[current_session_start]) * 100
                        short_session_changes.append(session_change)
            
            # Start new session
            current_session_start = i
            current_session_type = signal
            current_session_high = high[i]
            current_session_low = low[i]
        
        # Update session high/low
        if current_session_start is not None:
            if high[i] > current_session_high:
                current_session_high = high[i]
            if low[i] < current_session_low:
                current_session_low = low[i]
    
    # Close final session
    if current_session_start is not None and current_session_type is not None:
        if current_session_type == 1:  # Long session
            if current_session_high is not None:
                session_change = ((current_session_high - close[current_session_start]) / close[current_session_start]) * 100
                long_session_changes.append(session_change)
        else:  # Short session
            if current_session_low is not None:
                session_change = ((close[current_session_start] - current_session_low) / close[current_session_start]) * 100
                short_session_changes.append(session_change)
    
    # Calculate statistics
    avg_long = float(np.mean(long_session_changes)) if long_session_changes else 0.0
    max_long = float(np.max(long_session_changes)) if long_session_changes else 0.0
    avg_short = float(np.mean(short_session_changes)) if short_session_changes else 0.0
    max_short = float(np.max(short_session_changes)) if short_session_changes else 0.0
    
    return {
        'avg_long_session_pct': avg_long,
        'max_long_session_pct': max_long,
        'avg_short_session_pct': avg_short,
        'max_short_session_pct': max_short,
        'total_sessions': len(long_session_changes) + len(short_session_changes),
        'long_sessions': len(long_session_changes),
        'short_sessions': len(short_session_changes)
    }


def get_30min_candles(pair: str, start: datetime, end: datetime):
    """Get 30-minute candles for SuperTrend analysis with chunked requests."""
    url = f"{BASE_URL}/products/{pair}/candles"
    all_data = []
    
    # Calculate chunk size - Coinbase typically limits to ~300 candles per request
    # For 30-minute candles, that's about 6.25 days per chunk
    chunk_days = 6  # Conservative chunk size
    chunk_size = timedelta(days=chunk_days)
    
    current_start = start
    request_count = 0
    max_requests = 20  # Safety limit
    
    try:
        while current_start < end and request_count < max_requests:
            current_end = min(current_start + chunk_size, end)
            
            params = {
                "start": iso_format(current_start),
                "end": iso_format(current_end),
                "granularity": 1800,  # 30 minutes = 1800 seconds
            }
            
            resp = requests.get(url, headers=HEADERS, params=params, timeout=30)
            
            if resp.status_code == 400:
                # If 30-minute fails, try 1-hour granularity
                params["granularity"] = 3600  # 1 hour = 3600 seconds
                resp = requests.get(url, headers=HEADERS, params=params, timeout=30)
                if resp.status_code != 200:
                    print(f"Warning: {pair} doesn't support 30-minute or 1-hour granularity")
                    return []
            
            resp.raise_for_status()
            data = resp.json()

            if not isinstance(data, list):
                msg = data.get("message", "Unknown error format")
                raise RuntimeError(f"Candles error for {pair}: {msg}")

            all_data.extend(data)
            current_start = current_end
            request_count += 1
            
            # Rate limiting
            time.sleep(RATE_LIMIT_SLEEP)

        # Remove duplicates and sort
        unique_data = []
        seen_timestamps = set()
        for candle in all_data:
            timestamp = candle[0]
            if timestamp not in seen_timestamps:
                unique_data.append(candle)
                seen_timestamps.add(timestamp)
        
        unique_data.sort(key=lambda r: r[0])  # oldest -> newest
        return unique_data
        
    except requests.HTTPError as e:
        if "400" in str(e):
            print(f"Warning: {pair} API limitation - returning empty data")
            return []
        raise
    except Exception as e:
        print(f"Error fetching candles for {pair}: {e}")
        return []


def get_supertrend_stats(pair: str, start: datetime, end: datetime, factor=3, atr_length=10):
    """
    Get SuperTrend statistics for a trading pair.
    
    Returns:
        dict: SuperTrend session statistics
    """
    try:
        # Get 30-minute candles
        candles = get_30min_candles(pair, start, end)
        
        if len(candles) < atr_length + 10:  # Need minimum data for analysis
            return {
                'avg_long_session_pct': 0.0,
                'max_long_session_pct': 0.0,
                'avg_short_session_pct': 0.0,
                'max_short_session_pct': 0.0,
                'total_sessions': 0,
                'long_sessions': 0,
                'short_sessions': 0
            }
        
        # Extract OHLC data
        high = [float(row[2]) for row in candles]  # high is at index 2
        low = [float(row[3]) for row in candles]    # low is at index 3
        close = [float(row[4]) for row in candles] # close is at index 4
        
        # Calculate SuperTrend
        supertrend_line, trend_direction, signals = calculate_supertrend(
            high, low, close, factor, atr_length
        )
        
        if supertrend_line is None:
            return {
                'avg_long_session_pct': 0.0,
                'max_long_session_pct': 0.0,
                'avg_short_session_pct': 0.0,
                'max_short_session_pct': 0.0,
                'total_sessions': 0,
                'long_sessions': 0,
                'short_sessions': 0
            }
        
        # Analyze sessions
        session_stats = analyze_supertrend_sessions(
            high, low, close, supertrend_line, trend_direction, signals
        )
        
        return session_stats
        
    except Exception as e:
        print(f"Error calculating SuperTrend for {pair}: {e}")
        return {
            'avg_long_session_pct': 0.0,
            'max_long_session_pct': 0.0,
            'avg_short_session_pct': 0.0,
            'max_short_session_pct': 0.0,
            'total_sessions': 0,
            'long_sessions': 0,
            'short_sessions': 0
        }


# ----------------------------
# Coinbase API
# ----------------------------
def get_active_pairs(quote_currency: str = "USD"):
    url = f"{BASE_URL}/products"
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    products = resp.json()

    active = []
    # Convert to uppercase for consistent matching
    quote_currency_upper = quote_currency.upper()
    quote_suffix = f"-{quote_currency_upper}"
    
    for p in products:
        if p.get("trading_disabled"):
            continue
        if p.get("cancel_only"):
            continue
        if p.get("status") not in (None, "online", "active", "online_trading"):
            continue

        product_id = p.get("id") or p.get("product_id")
        if not product_id:
            continue

        if not product_id.endswith(quote_suffix):
            continue

        active.append(product_id)

    return sorted(set(active))


def get_pair_info(pair: str):
    url = f"{BASE_URL}/products/{pair}"
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    product_info = resp.json()
    return product_info.get("min_market_funds", None)


def get_daily_volume(pair: str, start: datetime, end: datetime):
    """Get the median daily volume over the specified period"""
    url = f"{BASE_URL}/products/{pair}/candles"
    params = {
        "start": iso_format(start),
        "end": iso_format(end),
        "granularity": GRANULARITY,
    }
    resp = requests.get(url, headers=HEADERS, params=params, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    if not isinstance(data, list):
        msg = data.get("message", "Unknown error format")
        raise RuntimeError(f"Candles error for {pair}: {msg}")

    if not data:
        return 0.0

    # Calculate median daily volume (volume is at index 5 in the candles data)
    volumes = [float(row[5]) for row in data if len(row) > 5]
    return float(np.median(volumes)) if volumes else 0.0


def get_daily_ohlc(pair: str, start: datetime, end: datetime):
    url = f"{BASE_URL}/products/{pair}/candles"
    params = {
        "start": iso_format(start),
        "end": iso_format(end),     # end at UTC midnight -> no partial candle
        "granularity": GRANULARITY,
    }
    resp = requests.get(url, headers=HEADERS, params=params, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    if not isinstance(data, list):
        msg = data.get("message", "Unknown error format")
        raise RuntimeError(f"Candles error for {pair}: {msg}")

    data.sort(key=lambda r: r[0])  # oldest -> newest
    return data


# ----------------------------
# Main
# ----------------------------
def main(volatility_threshold: float = 2.0, days: int = 90, output_file: str = "volatility.xlsx", volume_threshold: float = 1000000.0, output_format: str = "excel", quote_currency: str = "USD"):
    try:
        # Normalize quote currency to uppercase for consistency
        quote_currency = quote_currency.upper()
        
        # Calculate date window based on days parameter
        end_date = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        start_date = end_date - timedelta(days=days)  # inclusive start
        
        print(f"Window: {start_date.isoformat()} to {end_date.isoformat()} (UTC), {days} FULL daily candles expected.")
        print(f"Using volatility threshold: {volatility_threshold}%")
        print(f"Using volume threshold: {volume_threshold:,.0f}")
        print(f"Quote currency: {quote_currency}")
        print(f"Output format: {output_format.upper()}")
        print(f"Output file: {output_file}")
        print("📊 SuperTrend Analysis: 30-min candles, Factor 3, ATR Length 10")
        print("Press Ctrl+C at any time to cancel the operation, or type 'q' when prompted.\n")
        
        # Safely remove existing file if it exists
        if not safe_remove_file(output_file):
            print("❌ Cannot proceed without removing the existing file.")
            return

        active_pairs = get_active_pairs(quote_currency)
        print(f"Found {len(active_pairs)} active -{quote_currency} pairs.")

        # Initialize output file based on format
        if output_format == "excel":
            wb = create_excel_file(output_file)
        else:
            wb = None  # CSV mode

        # Create progress bar
        progress_bar = tqdm(
            total=len(active_pairs),
            desc="Processing pairs",
            unit="pair",
            bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]",
            position=0,
            leave=True,
            dynamic_ncols=True
        )

        for i, pair in enumerate(active_pairs, start=1):
            try:
                tqdm.write(f"[{i}/{len(active_pairs)}] Fetching {pair} ...")
                ohlc = get_daily_ohlc(pair, start_date, end_date)

                # Enforce at least the required number of FULL daily candles
                if len(ohlc) < days:
                    tqdm.write(f"  Skipping {pair}: only {len(ohlc)} full day(s) of history in the {days}-day window.")
                    progress_bar.update(1)
                    time.sleep(RATE_LIMIT_SLEEP)
                    continue

                # Compute per-day % range, then the median
                pct_changes = []
                for row in ohlc[-days:]:
                    _, low, high, *_ = row
                    pct_changes.append(calculate_percentage_change(float(low), float(high)))

                median_pct = float(np.median(pct_changes)) if pct_changes else 0.0

                # NEW: Exclude under-threshold medians
                if median_pct < volatility_threshold:
                    tqdm.write(f"  Skipping {pair}: median {median_pct:.4f}% < {volatility_threshold:.2f}% threshold.")
                    progress_bar.update(1)
                    time.sleep(RATE_LIMIT_SLEEP)
                    continue

                tqdm.write(f"  Median daily range% ({days}d): {median_pct:.4f}")

                # Get median daily volume
                try:
                    median_volume = get_daily_volume(pair, start_date, end_date)
                    tqdm.write(f"  Median daily volume: {median_volume:,.2f}")
                except Exception as e:
                    tqdm.write(f"  Warning: could not fetch volume for {pair}: {e}")
                    median_volume = 0.0

                # NEW: Exclude under-threshold volumes
                if median_volume < volume_threshold:
                    tqdm.write(f"  Skipping {pair}: median volume {median_volume:,.2f} < {volume_threshold:,.0f} threshold.")
                    progress_bar.update(1)
                    time.sleep(RATE_LIMIT_SLEEP)
                    continue

                # Get minimum market funds
                try:
                    min_mkt_funds = get_pair_info(pair)
                except Exception as e:
                    tqdm.write(f"  Warning: could not fetch min_market_funds for {pair}: {e}")
                    min_mkt_funds = "N/A"

                # Get SuperTrend analysis (only for qualifying pairs)
                tqdm.write(f"  Fetching 30-minute candles for SuperTrend analysis...")
                try:
                    supertrend_stats = get_supertrend_stats(pair, start_date, end_date)
                    if supertrend_stats['total_sessions'] > 0:
                        tqdm.write(f"  SuperTrend: {supertrend_stats['total_sessions']} sessions, "
                                  f"Avg Long: {supertrend_stats['avg_long_session_pct']:.2f}%, "
                                  f"Max Long: {supertrend_stats['max_long_session_pct']:.2f}%, "
                                  f"Avg Short: {supertrend_stats['avg_short_session_pct']:.2f}%, "
                                  f"Max Short: {supertrend_stats['max_short_session_pct']:.2f}%")
                    else:
                        tqdm.write(f"  SuperTrend: No sufficient data for analysis")
                except Exception as e:
                    tqdm.write(f"  SuperTrend: No data available for {pair}")
                    supertrend_stats = {
                        'avg_long_session_pct': 0.0,
                        'max_long_session_pct': 0.0,
                        'avg_short_session_pct': 0.0,
                        'max_short_session_pct': 0.0,
                        'total_sessions': 0,
                        'long_sessions': 0,
                        'short_sessions': 0
                    }

                # Save data based on format
                if output_format == "excel":
                    save_to_excel(pair, median_pct, median_volume, min_mkt_funds, supertrend_stats, wb)
                else:
                    if not save_to_csv(pair, median_pct, median_volume, min_mkt_funds, supertrend_stats, output_file):
                        tqdm.write("❌ Failed to save data. Operation cancelled.")
                        progress_bar.close()
                        return

                tqdm.write(f"  ✅ {pair} added to results")

            except requests.HTTPError as e:
                tqdm.write(f"HTTP error for {pair}: {e} | Response: {getattr(e, 'response', None)}")
            except Exception as e:
                tqdm.write(f"Error for {pair}: {e}")

            progress_bar.update(1)
            time.sleep(RATE_LIMIT_SLEEP)

        # Close progress bar
        progress_bar.close()
        
        # Finalize output based on format
        if output_format == "excel":
            # Sort and format Excel file
            sort_excel_by_volatility(wb)
            format_excel_file(wb)
            
            # Save Excel file
            def save_excel_operation():
                wb.save(output_file)
                return True
            
            if not safe_file_operation(save_excel_operation, output_file, "save Excel file"):
                print("❌ Failed to save Excel file.")
                return
        else:
            # Sort CSV file
            if not sort_csv_by_median(output_file):
                print("❌ Failed to sort the CSV file. Data may be saved but not sorted.")
                return
        
        print(f"\n✅ Done. Results saved (and sorted) in {output_file}")
        
        # Auto-open the file
        print("📂 Opening file...")
        if open_file(output_file):
            print("✅ File opened successfully!")
        else:
            print("ℹ️  File saved but could not be auto-opened.")
        
    except KeyboardInterrupt:
        print(f"\n\n🛑 Operation cancelled by user (Ctrl+C).")
        print("Partial data may have been saved to the output file.")
        if 'progress_bar' in locals():
            progress_bar.close()
        sys.exit(0)


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Fetch Coinbase coins with volatility above threshold over specified number of days"
    )
    parser.add_argument(
        "--volatility",
        type=float,
        default=2.0,
        help="Minimum daily median volatility percentage threshold (default: 2.0)"
    )
    parser.add_argument(
        "--days",
        type=int,
        default=90,
        help="Number of days to analyze for volatility (default: 90)"
    )
    parser.add_argument(
        "--output",
        type=str,
        default="volatility.xlsx",
        help="Output file name and path (default: volatility.xlsx)"
    )
    parser.add_argument(
        "--volume",
        type=float,
        default=1000000.0,
        help="Minimum median daily volume threshold (default: 1000000.0)"
    )
    parser.add_argument(
        "--format",
        choices=["excel", "csv"],
        default="excel",
        help="Output format: excel or csv (default: excel)"
    )
    parser.add_argument(
        "--quote",
        type=str,
        default="USD",
        help="Quote currency for trading pairs (e.g., USD, BTC, ETH, USDC, USDT) (default: USD)"
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_arguments()
    main(volatility_threshold=args.volatility, days=args.days, output_file=args.output, volume_threshold=args.volume, output_format=args.format, quote_currency=args.quote)
